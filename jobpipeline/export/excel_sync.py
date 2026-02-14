from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from jobpipeline.core.models import JobRecord

COLUMNS = [
    "Job ID", "Date Collected", "Last Seen", "Company", "Title", "Location", "Remote", "Link", "Source(s)",
    "Posted Date", "Fit Score", "Fit Grade", "Fit Notes", "Missing Must-have", "Status", "Notes"
]


def sync_excel(path: str, jobs: list[JobRecord]) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(p) if p.exists() else Workbook()
    ws = wb["Jobs"] if "Jobs" in wb.sheetnames else wb.active
    ws.title = "Jobs"
    if ws.max_row == 1 and ws["A1"].value is None:
        ws.append(COLUMNS)
    existing: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        existing[str(ws.cell(row, 1).value)] = row
    for i, h in enumerate(COLUMNS, 1):
        ws.cell(1, i).value = h
        ws.cell(1, i).font = Font(bold=True)
    for job in jobs:
        values = [job.job_id, job.collected_at, job.last_seen, job.company, job.title, job.location_text, job.remote_flag,
                  job.job_url, job.source_name, job.posted_date, job.fit_score, job.fit_grade, job.fit_notes,
                  ", ".join(job.missing_must_have)]
        if job.job_id in existing:
            r = existing[job.job_id]
            for col, val in enumerate(values, 1):
                ws.cell(r, col).value = val
        else:
            ws.append(values + ["New", ""])
            r = ws.max_row
        ws.cell(r, 8).hyperlink = job.job_url
        ws.cell(r, 8).style = "Hyperlink"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in "ABCDEFGHIJKLMNOP":
        ws.column_dimensions[col].width = 18
    wb.save(p)
